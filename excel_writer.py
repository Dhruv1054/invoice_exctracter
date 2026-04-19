# -*- coding: utf-8 -*-
"""
Excel generation — one file per invoice.

Each Excel contains:
  Sheet 1 "Invoice Table"  — raw table extracted from the PDF, as-is
  Sheet 2 "Invoice Info"   — header meta (invoice #, date, customer, etc.)
"""
import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_BG   = "2F5496"
HEADER_FG   = "FFFFFF"
ALT_ROW_BG  = "EEF2FF"   # light blue-grey for alternate data rows
TOTAL_BG    = "D9E1F2"   # slightly darker for rows that look like totals

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def generate_invoice_excel(
    table_data: dict,
    invoice_meta: dict,
    output_path: str,
    source_filename: str = "",
) -> str:
    """
    Write one Excel file for a single invoice.

    Parameters
    ----------
    table_data      : result of parser.extract_best_table()
    invoice_meta    : result of parser.extract_invoice_meta()
    output_path     : full path to write the .xlsx
    source_filename : original PDF filename (written into info sheet)
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    headers = table_data.get("headers", [])
    rows    = table_data.get("rows",    [])

    # Build table DataFrame
    if headers and rows:
        n = len(headers)
        padded = [
            (r + [""] * (n - len(r)))[:n] for r in rows
        ]
        df_table = pd.DataFrame(padded, columns=headers)
    elif rows:
        df_table = pd.DataFrame(rows)
    else:
        df_table = pd.DataFrame({"Note": ["No table found in this PDF"]})

    # Build info DataFrame
    info_rows = []
    if source_filename:
        info_rows.append({"Field": "Source File", "Value": source_filename})
    info_rows.append({"Field": "Extracted At",
                      "Value": datetime.now().strftime("%d/%m/%Y %H:%M:%S")})
    for field, value in invoice_meta.items():
        info_rows.append({"Field": field, "Value": value})
    df_info = pd.DataFrame(info_rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_table.to_excel(writer, sheet_name="Invoice Table", index=False)
        df_info.to_excel(writer,  sheet_name="Invoice Info",  index=False)

    _apply_formatting(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Formatting
# ---------------------------------------------------------------------------

def _apply_formatting(path: str):
    wb = load_workbook(path)

    ws_table = wb["Invoice Table"]
    _style_header_row(ws_table)
    _auto_column_width(ws_table)
    _alternate_row_colours(ws_table)
    _highlight_total_rows(ws_table)
    ws_table.freeze_panes = "A2"

    ws_info = wb["Invoice Info"]
    _style_header_row(ws_info)
    _auto_column_width(ws_info)

    wb.save(path)


def _style_header_row(ws):
    for cell in ws[1]:
        cell.font      = Font(color=HEADER_FG, bold=True)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[1].height = 30


def _auto_column_width(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)
        for cell in col:
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True,
                                       horizontal="left")


def _alternate_row_colours(ws):
    alt_fill = PatternFill("solid", fgColor=ALT_ROW_BG)
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
        if i % 2 == 0:
            for cell in row:
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = alt_fill


def _highlight_total_rows(ws):
    """Bold + tinted background for rows whose first cell contains 'total'."""
    total_fill = PatternFill("solid", fgColor=TOTAL_BG)
    total_font = Font(bold=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        first = str(row[0].value or "").strip().lower()
        if "total" in first or first == "":
            # Check if any cell in this row has "total" keyword
            row_text = " ".join(str(c.value or "") for c in row).lower()
            if "total" in row_text and first in ("total", ""):
                for cell in row:
                    cell.fill = total_fill
                    cell.font = total_font
